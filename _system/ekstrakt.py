#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Ekstrakt dopasowań — filtruje rekordy DOPASOWANE po zakresie Data Ksiegowania."""

import sys
import sqlite3
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from raport import _ensure_wyjatki_table


def _resolve_dirs() -> tuple[Path, Path]:
    """Zwraca (app_dir, base_dir) niezależnie od tego, czy exe jest w _system\ czy w folderze głównym."""
    if getattr(sys, "frozen", False):
        exe_dir = Path(sys.executable).parent
    else:
        exe_dir = Path(__file__).parent
    if exe_dir.name == "_system":
        return exe_dir, exe_dir.parent
    return exe_dir / "_system", exe_dir


APP_DIR, BASE_DIR = _resolve_dirs()
DB_PATH = APP_DIR / "faktury.db"
RAPORTY = BASE_DIR / "Raporty"

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")

SAP_LOAD_COLS = [
    "rok_obrotowy", "okres_sprawozdawczy", "rodzaj_dokumentu",
    "referencja", "data_dokumentu", "data_ksiegowania",
    "numer_dokumentu", "klucz_referencyjny", "data_dekl_podat", "plik_zrodlowy",
]
KSEF_LOAD_COLS = [
    "invoice_number", "invoice_type", "issue_date", "ksef_number",
    "net_amount", "gross_amount", "vat_amount", "buyer_value", "buyer_name",
    "typ_korekty", "p6", "p6_do", "data_wyst_fa_korygowanej",
    "nr_fa_korygowanej", "nr_ksef_fa_korygowanej", "plik_zrodlowy",
]
OUTPUT_COLS = {
    "klucz_laczenia":            "Klucz laczenia",
    "Status":                    "Status",
    "rok_obrotowy":              "Rok",
    "okres_sprawozdawczy":       "Okres",
    "rodzaj_dokumentu":          "Rodzaj dok.",
    "referencja":                "Referencja SAP",
    "data_dokumentu":            "Data dok. SAP",
    "data_ksiegowania":          "Data ksiegowania",
    "numer_dokumentu":           "Nr dok. SAP",
    "klucz_referencyjny":        "Klucz ref. SAP",
    "data_dekl_podat":           "Data dekl. podat.",
    "invoice_number":            "Nr faktury KSeF",
    "invoice_type":              "Typ faktury",
    "issue_date":                "Data wystawienia",
    "ksef_number":               "Nr KSeF",
    "net_amount":                "Kwota netto",
    "gross_amount":              "Kwota brutto",
    "vat_amount":                "VAT",
    "buyer_value":               "NIP nabywcy",
    "buyer_name":                "Nabywca",
    "typ_korekty":               "Typ korekty",
    "p6":                        "P_6",
    "p6_do":                     "P_6_Do",
    "data_wyst_fa_korygowanej":  "Data wyst. FA kor.",
    "nr_fa_korygowanej":         "Nr FA kor.",
    "nr_ksef_fa_korygowanej":    "Nr KSeF FA kor.",
    "plik_sap":                  "Plik SAP",
    "plik_ksef":                 "Plik KSeF",
}


def _parse_date(s: str) -> str:
    """Parsuje datę w formacie DD.MM.YYYY lub YYYY-MM-DD → YYYY-MM-DD."""
    s = s.strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    raise ValueError(f"Nieprawidlowy format daty: '{s}'. Uzyj DD.MM.YYYY lub YYYY-MM-DD.")


def _ask_date(prompt: str) -> str:
    while True:
        raw = input(prompt).strip()
        try:
            return _parse_date(raw)
        except ValueError as e:
            print(f"  BLAD: {e}")


def _reconcile_matched(df_sap: pd.DataFrame, df_ksef: pd.DataFrame) -> pd.DataFrame:
    """Zwraca DataFrame dopasowanych rekordów z oryginalnymi nazwami kolumn."""
    sap  = df_sap[SAP_LOAD_COLS].rename(columns={"plik_zrodlowy": "plik_sap"})
    ksef = df_ksef[KSEF_LOAD_COLS].rename(columns={"plik_zrodlowy": "plik_ksef"})

    rv_mask   = sap["rodzaj_dokumentu"] == "RV"
    merged_rv = sap[rv_mask].merge(
        ksef, left_on="klucz_referencyjny", right_on="invoice_number", how="inner"
    )
    merged_dr = sap[~rv_mask].merge(
        ksef, left_on="referencja", right_on="invoice_number", how="inner"
    )
    df = pd.concat([merged_rv, merged_dr], ignore_index=True)
    df["klucz_laczenia"] = df["klucz_referencyjny"].where(
        df["rodzaj_dokumentu"] == "RV", df["referencja"]
    )
    df["Status"] = "DOPASOWANE"
    return df


def _format_sheet(ws) -> None:
    max_col = ws.max_column
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=False)
    for idx in range(1, max_col + 1):
        header_len = len(str(ws.cell(row=1, column=idx).value or ""))
        ws.column_dimensions[get_column_letter(idx)].width = min(header_len + 6, 45)
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}1"


def main():
    print("=" * 62)
    print("  EKSTRAKT DOPASOWAN SAP vs KSeF — LYRECO")
    print(f"  Baza: {DB_PATH}")
    print("=" * 62)
    print()

    if not DB_PATH.exists():
        print(f"BLAD: Baza danych nie istnieje: {DB_PATH}")
        print("Uruchom najpierw 'Importuj.exe'.")
        input("\nNacisnij Enter, aby zamknac...")
        return

    print("Podaj zakres dat dla pola 'Data Ksiegowania'")
    print("(format: DD.MM.YYYY lub YYYY-MM-DD)")
    print()
    date_od = _ask_date("  Data OD: ")
    date_do = _ask_date("  Data DO: ")

    if date_od > date_do:
        print("\nBLAD: Data OD musi byc wczesniejsza lub rowna Dacie DO.")
        input("\nNacisnij Enter, aby zamknac...")
        return

    print()
    RAPORTY.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(DB_PATH)
    _ensure_wyjatki_table(conn)

    print("Wczytywanie SAP  ...", end=" ", flush=True)
    df_sap  = pd.read_sql("SELECT * FROM faktury_sap",  conn)
    print(f"{len(df_sap):,} wierszy")

    print("Wczytywanie KSeF ...", end=" ", flush=True)
    df_ksef = pd.read_sql("SELECT * FROM faktury_ksef", conn)
    print(f"{len(df_ksef):,} wierszy")

    df_wyjatki = pd.read_sql(
        """
        SELECT regula           AS "Regula",
               klucz_laczenia   AS "Klucz laczenia",
               numer_dokumentu  AS "Nr dok. SAP",
               invoice_number   AS "Nr faktury KSeF",
               rodzaj_dokumentu AS "Rodzaj dok.",
               referencja_sap   AS "Referencja SAP",
               data_dokumentu   AS "Data dok. SAP",
               data_ksiegowania AS "Data ksiegowania",
               data_wystawienia AS "Data wystawienia",
               nabywca          AS "Nabywca",
               kwota_brutto     AS "Kwota brutto",
               komentarz        AS "Komentarz",
               data_akceptacji  AS "Data akceptacji"
        FROM wyjatki_akceptacja ORDER BY regula, data_akceptacji
        """,
        conn,
    )
    conn.close()

    print("Laczenie danych  ...", end=" ", flush=True)
    df = _reconcile_matched(df_sap, df_ksef)
    print(f"OK — {len(df):,} dopasowanych lacznie")

    df_filtered = df[
        (df["data_ksiegowania"] >= date_od) &
        (df["data_ksiegowania"] <= date_do)
    ].copy()
    print(f"  Po filtrze ({date_od} — {date_do}): {len(df_filtered):,} rekordow")

    if df_filtered.empty:
        print("\nBrak rekordow w podanym zakresie dat.")
        input("\nNacisnij Enter, aby zamknac...")
        return

    int_cols = list(OUTPUT_COLS.keys())
    for col in int_cols:
        if col not in df_filtered.columns:
            df_filtered[col] = None
    df_out = (
        df_filtered[int_cols]
        .rename(columns=OUTPUT_COLS)
        .sort_values("Data ksiegowania", ignore_index=True)
    )

    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    od_str   = date_od.replace("-", "")
    do_str   = date_do.replace("-", "")
    out_path = RAPORTY / f"Ekstrakt_Dopasowane_{od_str}_{do_str}_{ts}.xlsx"

    print(f"\nZapis Excel: {out_path.name} ...", end=" ", flush=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_out.to_excel(writer, sheet_name="Dopasowane", index=False)
        df_wyjatki.to_excel(writer, sheet_name="Wyjatki_Zaakceptowane", index=False)
        _format_sheet(writer.book["Dopasowane"])
        _format_sheet(writer.book["Wyjatki_Zaakceptowane"])
    print("OK")

    print(f"\nGotowe!  ({len(df_out):,} wierszy)")
    print(f"  {out_path}")
    input("\nNacisnij Enter, aby zamknac...")


if __name__ == "__main__":
    try:
        main()
    except Exception:
        import traceback
        print("\n" + "=" * 62)
        print("  BLAD KRYTYCZNY:")
        traceback.print_exc()
        print("=" * 62)
        input("\nNacisnij Enter, aby zamknac...")
