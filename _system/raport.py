#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generuj raport Excel z rekoncyliacji SAP vs KSeF.

Logika złączenia:
  RV  → SAP.klucz_referencyjny = KSeF.invoice_number
  DR (i pozostałe) → SAP.referencja = KSeF.invoice_number

Plik wynikowy: ./Raporty/Rekoncyliacja_YYYYMMDD_HHMMSS.xlsx
"""

import sys
import sqlite3
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter


# ── Ścieżki ───────────────────────────────────────────────────────────────────
def _resolve_dirs() -> tuple[Path, Path]:
    """Zwraca (app_dir, base_dir) niezależnie od tego, czy exe jest w _system\ czy w folderze głównym."""
    if getattr(sys, "frozen", False):
        exe_dir = Path(sys.executable).parent
    else:
        exe_dir = Path(__file__).parent
    if exe_dir.name == "_system":
        return exe_dir, exe_dir.parent
    return exe_dir / "_system", exe_dir


APP_DIR, BASE_DIR = _resolve_dirs()
DB_PATH = APP_DIR / "faktury.db"
RAPORTY = BASE_DIR / "Raporty"


# ── Kolory statusów ───────────────────────────────────────────────────────────
STATUS_COLOR = {
    "DOPASOWANE": "C6EFCE",
    "TYLKO SAP":  "FFEB9C",
    "TYLKO KSeF": "FFC7CE",
}

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")


# ── Indeksy SQLite (przyspiesza kolejne zapytania) ────────────────────────────
def _ensure_indexes(conn: sqlite3.Connection) -> None:
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sap_klucz ON faktury_sap(klucz_referencyjny)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sap_ref   ON faktury_sap(referencja)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ksef_inv  ON faktury_ksef(invoice_number)")
    conn.commit()


# ── Wyjątki (akceptowane niezgodności) ─────────────────────────────────────────
# Kolumny opisowe zapisywane jako "zrzut" stanu pozycji w momencie akceptacji —
# dzięki temu zakładka "Zaakceptowane wyjątki" pokazuje pełny kontekst nawet
# po tym, jak dana pozycja zniknie z bieżącego raportu.
_WYJATKI_SNAPSHOT_COLS = {
    "rodzaj_dokumentu":  "Rodzaj dok.",
    "data_dokumentu":    "Data dok. SAP",
    "data_ksiegowania":  "Data ksiegowania",
    "referencja_sap":    "Referencja SAP",
    "data_wystawienia":  "Data wystawienia",
    "nabywca":           "Nabywca",
    "kwota_brutto":      "Kwota brutto",
}


def _norm_val(v) -> str:
    """Normalizuje wartosc komorki do porownywalnego stringa ('' dla pustych/NaN)."""
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except (TypeError, ValueError):
        pass
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none", "nat") else s


def _ensure_wyjatki_table(conn: sqlite3.Connection) -> None:
    """Tworzy tabelę wyjątków, jeśli baza powstała przed dodaniem tej funkcji,
    i dokłada (migruje) kolumny dodane w późniejszych wersjach."""
    conn.execute("""
        CREATE TABLE IF NOT EXISTS wyjatki_akceptacja (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            regula           TEXT NOT NULL,
            klucz_laczenia   TEXT NOT NULL DEFAULT '',
            numer_dokumentu  TEXT NOT NULL DEFAULT '',
            invoice_number   TEXT NOT NULL DEFAULT '',
            rodzaj_dokumentu TEXT NOT NULL DEFAULT '',
            data_dokumentu   TEXT NOT NULL DEFAULT '',
            data_ksiegowania TEXT NOT NULL DEFAULT '',
            referencja_sap   TEXT NOT NULL DEFAULT '',
            data_wystawienia TEXT NOT NULL DEFAULT '',
            nabywca          TEXT NOT NULL DEFAULT '',
            kwota_brutto     TEXT NOT NULL DEFAULT '',
            komentarz        TEXT NOT NULL,
            data_akceptacji  TEXT NOT NULL,
            UNIQUE(regula, klucz_laczenia, numer_dokumentu, invoice_number)
        )
    """)
    existing = {r[1] for r in conn.execute("PRAGMA table_info(wyjatki_akceptacja)").fetchall()}
    migration_cols = ["klucz_laczenia"] + list(_WYJATKI_SNAPSHOT_COLS.keys())
    for col in migration_cols:
        if col not in existing:
            conn.execute(f"ALTER TABLE wyjatki_akceptacja ADD COLUMN {col} TEXT NOT NULL DEFAULT ''")
    conn.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_wyjatki_key "
        "ON wyjatki_akceptacja(regula, klucz_laczenia, numer_dokumentu, invoice_number)"
    )
    conn.commit()


def _load_accepted(conn: sqlite3.Connection) -> dict[tuple[str, str, str, str], str]:
    """Zwraca {(regula, Klucz laczenia, Nr dok. SAP, Nr faktury KSeF): komentarz} dla aktywnych wyjątków."""
    rows = conn.execute(
        "SELECT regula, klucz_laczenia, numer_dokumentu, invoice_number, komentarz FROM wyjatki_akceptacja"
    ).fetchall()
    return {(r[0], r[1], r[2], r[3]): r[4] for r in rows}


def _keys_for(df: pd.DataFrame, regula: str) -> list[tuple[str, str, str, str]]:
    """Buduje listę kluczy (regula, Klucz laczenia, Nr dok. SAP, Nr faktury KSeF) dla wierszy df.

    Klucz laczenia jest jedynym polem unikatowym dla każdej pozycji — pozostałe dwa
    pola zostają w kluczu dla kompatybilności z wcześniej zapisanymi wyjątkami.
    """
    def col_or_empty(name: str) -> pd.Series:
        return df[name] if name in df.columns else pd.Series([""] * len(df), index=df.index)

    key_k = col_or_empty("Klucz laczenia").map(_norm_val)
    key_a = col_or_empty("Nr dok. SAP").map(_norm_val)
    key_b = col_or_empty("Nr faktury KSeF").map(_norm_val)
    return list(zip([regula] * len(df), key_k, key_a, key_b))


def _filter_accepted(df: pd.DataFrame, regula: str, accepted: dict) -> pd.DataFrame:
    """Usuwa z df wiersze, których klucz (regula, Nr dok. SAP, Nr faktury KSeF) jest w accepted."""
    if df.empty:
        return df
    keys = _keys_for(df, regula)
    mask = [k not in accepted for k in keys]
    return df[mask].reset_index(drop=True)


def _wyjatek_komentarz_col(df: pd.DataFrame, regula: str, accepted: dict) -> pd.Series:
    """Zwraca Series z komentarzem wyjątku dla wierszy zaakceptowanych, '' dla pozostałych."""
    if df.empty:
        return pd.Series([], dtype=str, index=df.index)
    keys = _keys_for(df, regula)
    return pd.Series([accepted.get(k, "") for k in keys], index=df.index)


# ── Rekoncyliacja w pandas ────────────────────────────────────────────────────
SAP_LOAD_COLS  = [
    "rok_obrotowy", "okres_sprawozdawczy", "rodzaj_dokumentu",
    "referencja", "data_dokumentu", "data_ksiegowania",
    "numer_dokumentu", "klucz_referencyjny", "data_dekl_podat", "plik_zrodlowy",
]
KSEF_LOAD_COLS = [
    "invoice_number", "invoice_type", "issue_date", "ksef_number",
    "net_amount", "gross_amount", "vat_amount", "buyer_value", "buyer_name",
    "typ_korekty", "p6", "p6_do", "data_wyst_fa_korygowanej",
    "nr_fa_korygowanej", "nr_ksef_fa_korygowanej", "plik_zrodlowy",
]

# Kolejność i nazwy wyświetlane w raporcie
OUTPUT_COLS = {
    "klucz_laczenia":            "Klucz laczenia",
    "Status":                    "Status",
    "rok_obrotowy":              "Rok",
    "okres_sprawozdawczy":       "Okres",
    "rodzaj_dokumentu":          "Rodzaj dok.",
    "referencja":                "Referencja SAP",
    "data_dokumentu":            "Data dok. SAP",
    "data_ksiegowania":          "Data ksiegowania",
    "numer_dokumentu":           "Nr dok. SAP",
    "klucz_referencyjny":        "Klucz ref. SAP",
    "data_dekl_podat":           "Data dekl. podat.",
    "invoice_number":            "Nr faktury KSeF",
    "invoice_type":              "Typ faktury",
    "issue_date":                "Data wystawienia",
    "ksef_number":               "Nr KSeF",
    "net_amount":                "Kwota netto",
    "gross_amount":              "Kwota brutto",
    "vat_amount":                "VAT",
    "buyer_value":               "NIP nabywcy",
    "buyer_name":                "Nabywca",
    "typ_korekty":               "Typ korekty",
    "p6":                        "P_6",
    "p6_do":                     "P_6_Do",
    "data_wyst_fa_korygowanej":  "Data wyst. FA kor.",
    "nr_fa_korygowanej":         "Nr FA kor.",
    "nr_ksef_fa_korygowanej":    "Nr KSeF FA kor.",
    "plik_sap":                  "Plik SAP",
    "plik_ksef":                 "Plik KSeF",
}


def _reconcile(
    df_sap: pd.DataFrame, df_ksef: pd.DataFrame
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Zwraca (df_raw_matched, df_display).

    df_raw_matched – dopasowane rekordy z oryginalnymi nazwami kolumn z bazy
                     (do obliczeń niezgodności).
    df_display     – wszystkie rekordy (DOPASOWANE + TYLKO SAP + TYLKO KSeF)
                     z polskimi nazwami kolumn (do arkuszy głównych).
    """
    sap  = df_sap[SAP_LOAD_COLS].rename(columns={"plik_zrodlowy": "plik_sap"})
    ksef = df_ksef[KSEF_LOAD_COLS].rename(columns={"plik_zrodlowy": "plik_ksef"})

    # RV → klucz_referencyjny,  DR (i pozostałe) → referencja
    rv_mask = sap["rodzaj_dokumentu"] == "RV"
    merged_rv = sap[rv_mask].merge(
        ksef, left_on="klucz_referencyjny", right_on="invoice_number", how="left"
    )
    merged_dr = sap[~rv_mask].merge(
        ksef, left_on="referencja", right_on="invoice_number", how="left"
    )
    df = pd.concat([merged_rv, merged_dr], ignore_index=True)

    df["klucz_laczenia"] = df["klucz_referencyjny"].where(
        df["rodzaj_dokumentu"] == "RV", df["referencja"]
    )
    df["Status"] = df["invoice_number"].notna().map({True: "DOPASOWANE", False: "TYLKO SAP"})

    # Surowe dane dopasowanych rekordów (oryginalne nazwy kolumn)
    df_raw_matched = df[df["Status"] == "DOPASOWANE"].copy()

    # KSeF bez dopasowania w SAP
    matched = set(df["invoice_number"].dropna())
    ksef_only = ksef[~ksef["invoice_number"].isin(matched)].copy()
    ksef_only["Status"]         = "TYLKO KSeF"
    ksef_only["klucz_laczenia"] = ksef_only["invoice_number"]

    int_cols = list(OUTPUT_COLS.keys())
    for col in int_cols:
        for frame in (df, ksef_only):
            if col not in frame.columns:
                frame[col] = None

    df_all = pd.concat([df[int_cols], ksef_only[int_cols]], ignore_index=True)
    df_all = df_all.sort_values(["Status", "klucz_laczenia"], ignore_index=True)
    df_display = df_all.rename(columns=OUTPUT_COLS)

    return df_raw_matched, df_display


# ── Niezgodności ─────────────────────────────────────────────────────────────

def _empty(s: pd.Series) -> pd.Series:
    """True jeśli wartość jest None/NaN/pusty string."""
    return s.isna() | (s.astype(str).str.strip().isin(["", "nan", "None", "NaT"]))


def _neq(a: pd.Series, b: pd.Series) -> pd.Series:
    """Różne wartości — ignoruje pary gdzie którakolwiek strona jest pusta."""
    return (~_empty(a)) & (~_empty(b)) & (a.astype(str) != b.astype(str))


def _disc_cols(df: pd.DataFrame) -> pd.DataFrame:
    int_cols = list(OUTPUT_COLS.keys())
    cols = [c for c in int_cols if c in df.columns]
    return df[cols].rename(columns=OUTPUT_COLS).reset_index(drop=True)


def _p6_eff(df: pd.DataFrame) -> pd.Series:
    """Zwraca p6_do jeśli wypełnione, w przeciwnym razie p6."""
    result = df["p6_do"].copy()
    result[_empty(df["p6_do"])] = df["p6"][_empty(df["p6_do"])]
    return result


def _to_dt(s: pd.Series) -> pd.Series:
    """Konwertuje kolumnę stringów YYYY-MM-DD na datetime (NaT dla pustych/błędnych)."""
    return pd.to_datetime(s, errors="coerce")


def _discrepancies(df_raw: pd.DataFrame, df_sap_all: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """7 arkuszy niezgodności — tylko rekordy DOPASOWANE z oryginalnymi nazwami kolumn."""

    m = df_raw  # już tylko dopasowane

    # ── Reguła 1 ─────────────────────────────────────────────────────────────
    # data_dokumentu SAP ≠ issue_date KSeF
    r1 = m[_neq(m["data_dokumentu"], m["issue_date"])].copy()

    # Wykluczenie symetryczne: RV i powiązany z nim DR wzajemnie się znoszą,
    # jeśli spełnione są łącznie:
    #   klucz_referencyjny RV = referencja DR
    #   data_dokumentu RV = data_dokumentu DR
    #   data_dekl_podat RV = data_dekl_podat DR
    dr_all = df_sap_all[df_sap_all["rodzaj_dokumentu"] == "DR"]
    rv_all = df_sap_all[df_sap_all["rodzaj_dokumentu"] == "RV"]

    # Zestaw kluczy DR: (referencja, data_dokumentu, data_dekl_podat)
    dr_keys = set(zip(
        dr_all["referencja"].astype(str),
        dr_all["data_dokumentu"].astype(str),
        dr_all["data_dekl_podat"].astype(str),
    ))
    # Zestaw kluczy RV: (klucz_referencyjny, data_dokumentu, data_dekl_podat)
    rv_keys = set(zip(
        rv_all["klucz_referencyjny"].astype(str),
        rv_all["data_dokumentu"].astype(str),
        rv_all["data_dekl_podat"].astype(str),
    ))

    def _wyklucz(row) -> bool:
        if row["rodzaj_dokumentu"] == "RV":
            # Wyklucz RV gdy istnieje DR z tymi samymi datami i referencją
            return (str(row["klucz_referencyjny"]), str(row["data_dokumentu"]), str(row["data_dekl_podat"])) in dr_keys
        if row["rodzaj_dokumentu"] == "DR":
            # Wyklucz DR gdy istnieje RV z tymi samymi datami i kluczem
            return (str(row["referencja"]), str(row["data_dokumentu"]), str(row["data_dekl_podat"])) in rv_keys
        return False

    excl_mask = pd.Series([_wyklucz(row) for _, row in r1.iterrows()], index=r1.index)
    r1 = r1[~excl_mask]

    # ── Reguła 2 ─────────────────────────────────────────────────────────────
    # invoice_type = Vat: data_dekl_podat ≠ p6_do (lub p6, gdy p6_do puste)
    vat = m[m["invoice_type"] == "Vat"]
    has_p6do = ~_empty(vat["p6_do"])
    r2 = vat[
        (has_p6do  & _neq(vat["data_dekl_podat"], vat["p6_do"].astype(str))) |
        (~has_p6do & _neq(vat["data_dekl_podat"], vat["p6"]))
    ]

    # ── Reguła 3 ─────────────────────────────────────────────────────────────
    # invoice_type = Kor, typ_korekty = 1: data_dekl_podat ≠ p6_do (lub p6, gdy p6_do puste)
    kor1 = m[(m["invoice_type"] == "Kor") & (m["typ_korekty"] == "1")]
    has_p6do_k1 = ~_empty(kor1["p6_do"])
    r3 = kor1[
        (has_p6do_k1  & _neq(kor1["data_dekl_podat"], kor1["p6_do"].astype(str))) |
        (~has_p6do_k1 & _neq(kor1["data_dekl_podat"], kor1["p6"]))
    ]

    # ── Reguła 4 ─────────────────────────────────────────────────────────────
    # invoice_type = Kor, typ_korekty = 2: data_dekl_podat ≠ issue_date LUB ≠ data_dokumentu
    kor2 = m[(m["invoice_type"] == "Kor") & (m["typ_korekty"] == "2")]
    r4 = kor2[
        _neq(kor2["data_dekl_podat"], kor2["issue_date"]) |
        _neq(kor2["data_dekl_podat"], kor2["data_dokumentu"])
    ]

    # ── Reguła 5 ─────────────────────────────────────────────────────────────
    # invoice_type = Kor, data_wyst_fa_korygowanej > 2026-02-01:
    # nr_fa_korygowanej LUB nr_ksef_fa_korygowanej jest puste
    kor = m[m["invoice_type"] == "Kor"]
    mask5 = (
        (~_empty(kor["data_wyst_fa_korygowanej"])) &
        (kor["data_wyst_fa_korygowanej"] > "2026-02-01") &
        (_empty(kor["nr_fa_korygowanej"]) | _empty(kor["nr_ksef_fa_korygowanej"]))
    )
    r5 = kor[mask5]

    # ── Reguła 6 ─────────────────────────────────────────────────────────────
    # invoice_type = Vat: data_dokumentu SAP jest o więcej niż 2 miesiące późniejsza niż P6_do (lub P6)
    vat6   = m[m["invoice_type"] == "Vat"]
    p6_dt  = _to_dt(_p6_eff(vat6))
    doc_dt = _to_dt(vat6["data_dokumentu"])
    valid6 = doc_dt.notna() & p6_dt.notna()
    diff_months = (doc_dt.dt.year * 12 + doc_dt.dt.month) - (p6_dt.dt.year * 12 + p6_dt.dt.month)
    r6 = vat6[valid6 & (diff_months > 2)]

    # ── Reguła 7 ─────────────────────────────────────────────────────────────
    # invoice_type = Vat: P6_do (lub P6) jest w późniejszym miesiącu niż data_dokumentu SAP
    vat7   = m[m["invoice_type"] == "Vat"]
    p6_dt7 = _to_dt(_p6_eff(vat7))
    doc_dt7 = _to_dt(vat7["data_dokumentu"])
    valid7 = doc_dt7.notna() & p6_dt7.notna()
    p6_ym  = p6_dt7.dt.year  * 12 + p6_dt7.dt.month
    doc_ym = doc_dt7.dt.year * 12 + doc_dt7.dt.month
    r7 = vat7[valid7 & (p6_ym > doc_ym)]

    return {
        "Niezg_1_Daty":        _disc_cols(r1),
        "Niezg_2_VAT_Okres":   _disc_cols(r2),
        "Niezg_3_KOR1_Daty":   _disc_cols(r3),
        "Niezg_4_KOR2_Daty":   _disc_cols(r4),
        "Niezg_5_KOR_BrakNr":  _disc_cols(r5),
        "Niezg_6_Daty_2M":     _disc_cols(r6),
        "Niezg_7_Pozny_Okres": _disc_cols(r7),
    }


def _to_num(s: pd.Series) -> pd.Series:
    """Konwertuje kolumnę na liczby (NaN dla pustych/błędnych)."""
    return pd.to_numeric(s, errors="coerce")


def _korekta_faktura_ta_sama_wartosc(df_ksef: pd.DataFrame) -> pd.DataFrame:
    """Reguła 8 — Korekta_Faktura Ta Sama Wartość.

    Dokumenty o różnym invoice_type (Vat / Kor), tym samym issue_date i tym
    samym buyer_value, gdzie vat_amount jest taki sam co do wartości
    absolutnej (pomijając znak +/-).
    """
    ksef = df_ksef[KSEF_LOAD_COLS].rename(columns={"plik_zrodlowy": "plik_ksef"})
    ksef = ksef[ksef["invoice_type"].isin(["Vat", "Kor"])].copy()
    ksef["_abs_vat"] = _to_num(ksef["vat_amount"]).abs()

    valid = ksef[
        (~_empty(ksef["issue_date"])) &
        (~_empty(ksef["buyer_value"])) &
        ksef["_abs_vat"].notna()
    ]

    matched_idx = set()
    for _, group in valid.groupby(["issue_date", "buyer_value", "_abs_vat"]):
        types = set(group["invoice_type"])
        if "Vat" in types and "Kor" in types:
            matched_idx.update(group.index)

    result = ksef.loc[sorted(matched_idx)].drop(columns=["_abs_vat"])
    result = result.sort_values(["issue_date", "buyer_value", "invoice_type"], ignore_index=True)
    # Dla wierszy czysto-KSeF jedynym naturalnym kluczem unikatowym jest invoice_number
    # (tak jak dla Tylko_KSeF w _reconcile) — uzupelniamy klucz_laczenia dla spójności.
    result["klucz_laczenia"] = result["invoice_number"]
    return _disc_cols(result)


def _tylko_ksef_zero_z_p6(df_ksef_only: pd.DataFrame) -> pd.DataFrame:
    """Reguła 9 — Tylko KSeF: Kwota netto = Kwota brutto = VAT = 0, ale P_6_Do lub P_6 wypełnione."""
    netto  = _to_num(df_ksef_only["Kwota netto"])
    brutto = _to_num(df_ksef_only["Kwota brutto"])
    vat    = _to_num(df_ksef_only["VAT"])
    ma_p6  = (~_empty(df_ksef_only["P_6_Do"])) | (~_empty(df_ksef_only["P_6"]))
    mask   = (netto == 0) & (brutto == 0) & (vat == 0) & ma_p6
    return df_ksef_only[mask].reset_index(drop=True)


# ── Formatowanie arkusza ──────────────────────────────────────────────────────
def _format_sheet(ws, status_col_idx: int | None = None) -> None:
    """Nagłówek + szerokości kolumn + filtr. Kolorowanie przez ConditionalFormatting."""
    max_col = ws.max_column
    max_row = ws.max_row

    # Nagłówek
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=False)

    # ConditionalFormatting — Excel koloruje po otwarciu pliku (bez iteracji po wierszach)
    if status_col_idx and max_row > 1:
        col_letter = get_column_letter(status_col_idx)
        rng = f"A2:{get_column_letter(max_col)}{max_row}"
        for status, color in STATUS_COLOR.items():
            fill    = PatternFill("solid", fgColor=color)
            formula = f'${col_letter}2="{status}"'
            ws.conditional_formatting.add(rng, FormulaRule(formula=[formula], fill=fill))

    # Szerokości kolumn — tylko z nagłówka (unikamy iteracji przez dziesiątki tys. wierszy)
    for idx in range(1, max_col + 1):
        header_len = len(str(ws.cell(row=1, column=idx).value or ""))
        ws.column_dimensions[get_column_letter(idx)].width = min(header_len + 6, 45)

    ws.freeze_panes = "A2"
    if max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}1"


# ── Podsumowanie ──────────────────────────────────────────────────────────────
def _make_summary(conn: sqlite3.Connection, df_all: pd.DataFrame) -> pd.DataFrame:
    sap_total  = conn.execute("SELECT COUNT(*) FROM faktury_sap").fetchone()[0]
    ksef_total = conn.execute("SELECT COUNT(*) FROM faktury_ksef").fetchone()[0]
    counts     = df_all["Status"].value_counts()

    rows = [
        {"Metryka": "Lacznie faktur SAP",       "Wartosc": sap_total},
        {"Metryka": "Lacznie faktur KSeF",       "Wartosc": ksef_total},
        {"Metryka": "---",                        "Wartosc": "---"},
        {"Metryka": "Dopasowane",                "Wartosc": counts.get("DOPASOWANE", 0)},
        {"Metryka": "Tylko SAP (brak w KSeF)",   "Wartosc": counts.get("TYLKO SAP",  0)},
        {"Metryka": "Tylko KSeF (brak w SAP)",   "Wartosc": counts.get("TYLKO KSeF", 0)},
        {"Metryka": "---",                        "Wartosc": "---"},
        {"Metryka": "Raport wygenerowany",        "Wartosc": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
    ]
    return pd.DataFrame(rows)


# ── Log importu ───────────────────────────────────────────────────────────────
SQL_LOG = """
SELECT typ AS "Typ", plik AS "Plik", data_importu AS "Data importu",
       wiersze_dodane AS "Dodano", wiersze_pominiete AS "Pominieto", status AS "Status"
FROM import_log ORDER BY data_importu DESC LIMIT 1000
"""


def compute_report_data(
    conn: sqlite3.Connection,
) -> tuple[pd.DataFrame, dict[str, pd.DataFrame], pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Wczytuje dane i liczy rekoncyliację.

    Zwraca (df_all, disc, df_log, df_sap_full, df_ksef_full):
      df_all      – wszystkie rekordy (DOPASOWANE + TYLKO SAP + TYLKO KSeF), nieprzefiltrowane.
      disc        – 11 kategorii (Niezg_1..Niezg_9, Tylko_SAP, Tylko_KSeF), już odfiltrowane
                    o zaakceptowane wyjątki — jednolity zbiór do raportu i do Wyjatki.exe.
      df_log      – log importu.
      df_sap_full, df_ksef_full – pełne (nieprzefiltrowane) "Tylko SAP"/"Tylko KSeF",
                    do głównych zakładek inwentaryzacyjnych raportu.
    """
    _ensure_wyjatki_table(conn)

    df_sap  = pd.read_sql("SELECT * FROM faktury_sap",  conn)
    df_ksef = pd.read_sql("SELECT * FROM faktury_ksef", conn)
    df_log  = pd.read_sql(SQL_LOG, conn)

    df_raw, df_all = _reconcile(df_sap, df_ksef)

    disc = _discrepancies(df_raw, df_sap)
    disc["Niezg_8_Kor_Ta_Sama_Wartosc"] = _korekta_faktura_ta_sama_wartosc(df_ksef)

    df_sap_full  = df_all[df_all["Status"] == "TYLKO SAP"].reset_index(drop=True)
    df_ksef_full = df_all[df_all["Status"] == "TYLKO KSeF"].reset_index(drop=True)

    disc["Niezg_9_TylkoKSeF_ZeroP6"] = _tylko_ksef_zero_z_p6(df_ksef_full)
    disc["Tylko_SAP"]  = df_sap_full
    disc["Tylko_KSeF"] = df_ksef_full

    accepted = _load_accepted(conn)
    disc = {name: _filter_accepted(df_d, name, accepted) for name, df_d in disc.items()}

    return df_all, disc, df_log, df_sap_full, df_ksef_full


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("=" * 62)
    print("  RAPORT REKONCYLIACJI SAP vs KSeF — LYRECO")
    print(f"  Baza: {DB_PATH}")
    print("=" * 62)

    if not DB_PATH.exists():
        print(f"\nBLAD: Baza danych nie istnieje: {DB_PATH}")
        print("Uruchom najpierw 'Importuj.exe'.")
        input("\nNacisnij Enter, aby zamknac...")
        return

    RAPORTY.mkdir(parents=True, exist_ok=True)
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = RAPORTY / f"Rekoncyliacja_{ts}.xlsx"

    conn = sqlite3.connect(DB_PATH)
    _ensure_indexes(conn)

    print("\nWczytywanie i laczenie danych ...", end=" ", flush=True)
    df_all, disc, df_log, df_sap_full, df_ksef_full = compute_report_data(conn)
    print("OK")

    accepted   = _load_accepted(conn)
    df_sum     = _make_summary(conn, df_all)
    df_wyjatki = pd.read_sql(
        """
        SELECT regula           AS "Regula",
               klucz_laczenia   AS "Klucz laczenia",
               numer_dokumentu  AS "Nr dok. SAP",
               invoice_number   AS "Nr faktury KSeF",
               rodzaj_dokumentu AS "Rodzaj dok.",
               referencja_sap   AS "Referencja SAP",
               data_dokumentu   AS "Data dok. SAP",
               data_ksiegowania AS "Data ksiegowania",
               data_wystawienia AS "Data wystawienia",
               nabywca          AS "Nabywca",
               kwota_brutto     AS "Kwota brutto",
               komentarz        AS "Komentarz",
               data_akceptacji  AS "Data akceptacji"
        FROM wyjatki_akceptacja ORDER BY regula, data_akceptacji
        """,
        conn,
    )
    conn.close()

    df_match = df_all[df_all["Status"] == "DOPASOWANE"]

    # Zakładki Tylko_SAP / Tylko_KSeF w głównym raporcie pokazują WSZYSTKIE rekordy
    # (inwentaryzacja) — z dodatkową kolumną komentarza dla pozycji już zaakceptowanych.
    df_sap_  = df_sap_full.copy()
    df_ksef_ = df_ksef_full.copy()
    df_sap_["Wyjatek_Komentarz"]  = _wyjatek_komentarz_col(df_sap_full,  "Tylko_SAP",  accepted)
    df_ksef_["Wyjatek_Komentarz"] = _wyjatek_komentarz_col(df_ksef_full, "Tylko_KSeF", accepted)

    # Niezg_X w disc to wersje już odfiltrowane o zaakceptowane wyjątki (do excela).
    disc_report = {k: v for k, v in disc.items() if k not in ("Tylko_SAP", "Tylko_KSeF")}

    print(f"  Dopasowane : {len(df_match):>6,}")
    print(f"  Tylko SAP  : {len(df_sap_):>6,}")
    print(f"  Tylko KSeF : {len(df_ksef_):>6,}")
    for name, df_d in disc_report.items():
        print(f"  {name:<22}: {len(df_d):>6,}")

    status_col = list(df_all.columns).index("Status") + 1  # 1-indexed

    print(f"\nZapis Excel + formatowanie: {out_path.name} ...", end=" ", flush=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_sum.to_excel(writer,     sheet_name="Podsumowanie", index=False)
        df_sap_.to_excel(writer,    sheet_name="Tylko_SAP",    index=False)
        df_ksef_.to_excel(writer,   sheet_name="Tylko_KSeF",   index=False)
        df_log.to_excel(writer,     sheet_name="Log_importu",  index=False)
        df_wyjatki.to_excel(writer, sheet_name="Wyjatki_Zaakceptowane", index=False)
        for sheet_name, df_d in disc_report.items():
            df_d.to_excel(writer, sheet_name=sheet_name, index=False)

        wb = writer.book
        for name in ("Tylko_SAP", "Tylko_KSeF"):
            _format_sheet(wb[name], status_col)
        for name in ("Podsumowanie", "Log_importu", "Wyjatki_Zaakceptowane"):
            _format_sheet(wb[name])
        for name in disc_report:
            _format_sheet(wb[name])
    print("OK")

    print(f"\nGotowe!")
    print(f"  {out_path}")
    input("\nNacisnij Enter, aby zamknac...")


if __name__ == "__main__":
    try:
        main()
    except Exception:
        import traceback
        print("\n" + "=" * 62)
        print("  BLAD KRYTYCZNY:")
        traceback.print_exc()
        print("=" * 62)
        input("\nNacisnij Enter, aby zamknac...")
